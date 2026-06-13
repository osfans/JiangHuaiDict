#!/usr/bin/env python3
from pathlib import Path

from openpyxl import load_workbook
from collections import defaultdict
from opencc import OpenCC
cc = OpenCC("t2s")


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "淮語詞典230301initial.xlsx"
OUTPUT_MD = ROOT / "000词表.md"


def as_text(value) -> str:
    if value is None:
        return ""
    return cc.convert(str(value).strip())


def build_wordlist() -> list[str]:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.worksheets[1]

    max_col = ws.max_column
    max_row = ws.max_row

    # 从第1行读取地名表头，C列及以后为各地解释列
    place_headers = {
        col: as_text(ws.cell(row=1, column=col).value).replace("（府城）", "").replace("（", "(").replace("）", ")") for col in range(3, max_col + 1)
    }
    place_count = defaultdict(int)

    lines: list[str] = [f"# {SOURCE_XLSX.name}", ""]

    isLiangci = False
    for row in range(2, max_row + 1):
        head = as_text(ws.cell(row=row, column=2).value)
        if not head:
            continue
        if head == "量词":
            isLiangci = True
        elif head == "数词和相关用词":
            isLiangci = False
        if isLiangci:
            head = f"{head}(量词)"

        merged_parts: list[tuple[list[str], str]] = []
        for col in range(3, max_col + 1):
            explanation = as_text(ws.cell(row=row, column=col).value)
            if not explanation or explanation == "/":
                continue
            place = place_headers.get(col, "") or f"第{col}列"
            place_count[place] += 1

            # 若当前列解释与前某一列一致，则合并地名标签
            if merged_parts:
                for i, (_, exp) in enumerate(merged_parts):
                    if exp == explanation:
                        merged_parts[i][0].append(place)
                        break
                else:
                    merged_parts.append(([place], explanation))
            else:
                merged_parts.append(([place], explanation))

        if not merged_parts:
            # 无任何解释则跳过
            continue

        parts = [f"{''.join(f'〔{p}〕' for p in places)}{exp}" for places, exp in merged_parts]
        lines.append(f"【{head}】{''.join(parts)}  ")
    lines[1] = f"{''.join(f'〔{place}〕{count}' for place, count in place_count.items())}"

    return lines


def main() -> None:
    lines = build_wordlist()
    OUTPUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"已生成: {OUTPUT_MD}")
    print(f"词条数: {max(len(lines) - 2, 0)}")


if __name__ == "__main__":
    main()
